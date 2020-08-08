import requests
from bs4 import BeautifulSoup as BS
from openpyxl import load_workbook

SITE_COL = 3
PRODUCT_COL = 4
PRICE_COL = 5
START_ROW = 4

products_workbook = load_workbook('products.xlsx')
product_worksheet = products_workbook['products']

r = requests.get('https://shine-bright.com.ua/balibody')
html = BS(r.content, 'html.parser')
# print(html)
product_worksheet.cell(column=SITE_COL, row=START_ROW, value='Shine Bright')

for idx, product in enumerate(html.select('.product')):
    # print(el)
    name = product.select('.data-gtag-name')
    price = product.select('.data-gtag-price')
    print(name[0].text, price[0].text)
    product_worksheet.cell(column=PRODUCT_COL, row=START_ROW + idx, value=name[0].text)
    product_worksheet.cell(column=PRICE_COL, row=START_ROW + idx, value=price[0].text)

products_workbook.save(filename='products.xlsx')
